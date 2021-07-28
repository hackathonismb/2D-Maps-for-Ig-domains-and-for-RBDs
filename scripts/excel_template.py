from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import json
from colour import Color
from sys import argv
from copy import copy

colorcoded = int(argv[1])


if colorcoded:
	proteins = ["5ESV"]
	mapper = {"5ESV":{"light":"2","heavy":"1_contacts"}}
else:
	proteins = ["1RHH","2IG2","2ZJS","3DGG","4ZSO","5CMA","5ESV"]
	mapper = {	"1RHH":{"light":"1","heavy":"2"},
				"2IG2":{"light":"1","heavy":"2"},
				"2ZJS":{"light":"4","heavy":"3"},
				"3DGG":{"light":"1","heavy":"2"},
				"4ZSO":{"light":"1","heavy":"2"},
				"5CMA":{"light":"1","heavy":"2"},
				"5ESV":{"light":"2","heavy":"1"}}
col_range = {
	"heavy":['D','E','F','G','H','I','J','K','L','M','N','O','P'],
	"light":['X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']
	}
row_range = {
	"heavy":[str(i) for i in range(10,42)],
	"light":[str(i) for i in range(10,31)]
	}
properties = ["contacts"]
colors = [Color("black")]+list(Color("blue").range_to(Color("green"),15))
properties_mapper = {"contacts":{i:"FF"+colors[i].hex_l[1:].upper() for i in range(len(colors))}}


#for each protein
for protein in proteins:
	print(protein)
	if colorcoded:
		wb = load_workbook("igv_hl_template.xlsx")
	else:
		wb = load_workbook("igv_hl_template_default_colours.xlsx")
	ws = wb.active

#for both heavy and light chain
	for chain in ["heavy","light"]:
		print(mapper[protein][chain])
		json_file = open("proteins/"+protein+"_"+mapper[protein][chain]+".json")
		data = json.load(json_file)

#for each residue
		for res in data["residues"]:
			if "kabat" not in res.keys():
				continue
			kabat = res["kabat"]

#search the template for the correct position
			for col in col_range[chain]:
				for row in row_range[chain]:
					cell = col+row
					if str(ws[cell].value) == kabat:

#fill the cell with the information
						ws[cell].value = res["res"]
						for p in properties:
							if p in res.keys():
								pv = res[p]
								#CHANGE ws[cell] ACCORDINGLY
								f = copy(ws[cell].font)
								#f.b = True #switch to bold
								f.color = properties_mapper[p][pv] #switch to black text
								#print(f)
								ws[cell].font = f #apply changes on cell
								#print(ws[cell].font)
								#assert(False)
								#f = copy(ws[cell].fill)
								#f.fgColor = "FF000000" #switch background color to black
								#ws[cell].fill = f #apply changes on cell

#clean numbering without corresponding position
		for col in col_range[chain]:
			for row in row_range[chain]:
				cell = col+row
				if str(ws[cell].value)[0] in [str(i) for i in range(10)]:
					ws[cell].value = ""
#save the protein map
		if colorcoded:
			wb.save("output2/"+protein+".xlsx")
		else:
			wb.save("output/"+protein+".xlsx")
